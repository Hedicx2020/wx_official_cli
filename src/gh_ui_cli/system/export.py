"""/api/export/excel 本地实现 - 用 openpyxl 写 xlsx。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from ..wechat.errors import WechatError, WechatInvalidInput
from ..wechat.registry import capability
from . import paths


def export_excel(payload: dict) -> dict[str, Any]:
    data = payload.get("data")
    columns = payload.get("columns")
    filename = (payload.get("filename") or "export").strip() or "export"
    sheet_name = (payload.get("sheet_name") or "data").strip() or "data"

    if not isinstance(data, list):
        raise WechatInvalidInput("data 必须是数组")
    if not isinstance(columns, list) or not columns:
        raise WechatInvalidInput("columns 必须是非空数组")

    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise WechatError(
            "缺少 openpyxl，请 uv pip install openpyxl",
            code="EXPORT_MISSING_DEP",
        ) from e

    out_dir = Path(paths.export_path())
    out_dir.mkdir(parents=True, exist_ok=True)
    filepath = out_dir / f"{filename}.xlsx"
    if filepath.exists():
        i = 1
        while (out_dir / f"{filename}_{i}.xlsx").exists():
            i += 1
        filepath = out_dir / f"{filename}_{i}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([str(c) for c in columns])
    for row in data:
        if isinstance(row, dict):
            ws.append([row.get(c) for c in columns])
        elif isinstance(row, (list, tuple)):
            ws.append(list(row))
    ws.freeze_panes = "A2"
    for idx, col in enumerate(columns, start=1):
        widths = [len(str(col))]
        for r in range(2, min(102, ws.max_row + 1)):
            widths.append(len(str(ws.cell(r, idx).value or "")))
        col_letter = ws.cell(1, idx).column_letter
        ws.column_dimensions[col_letter].width = min(max(widths) + 3, 40)

    wb.save(str(filepath))
    return {"path": str(filepath)}


@capability("op:system:export-excel")
def _cap(payload: dict) -> dict:
    return export_excel(payload or {})
